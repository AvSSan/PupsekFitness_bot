from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook

from kbju_bot.calculations import build_daily_summary
from kbju_bot.models import MealEntry, UserSettings
from kbju_bot.formatting import format_number


def build_excel_export(entries: list[MealEntry], settings: UserSettings | None) -> BytesIO:
    workbook = Workbook()
    entries_sheet = workbook.active
    entries_sheet.title = "entries"
    entries_sheet.append(["date", "meal", "description", "calories", "protein", "fat", "carbs", "created_at"])
    for entry in entries:
        entries_sheet.append(
            [
                entry.entry_date.isoformat(),
                entry.meal_name_raw,
                entry.description,
                entry.calories,
                entry.protein,
                entry.fat,
                entry.carbs,
                entry.created_at.isoformat(timespec="seconds"),
            ]
        )

    summary_sheet = workbook.create_sheet("daily_summary")
    summary_sheet.append(
        [
            "date",
            "total_calories",
            "total_protein",
            "total_fat",
            "total_carbs",
            "target_calories",
            "target_protein",
            "target_fat",
            "target_carbs",
            "remaining_calories",
            "remaining_protein",
            "remaining_fat",
            "remaining_carbs",
        ]
    )

    entries_by_date: dict[date, list[MealEntry]] = defaultdict(list)
    for entry in entries:
        entries_by_date[entry.entry_date].append(entry)

    for day in sorted(entries_by_date):
        summary = build_daily_summary(day, entries_by_date[day], settings)
        target = summary.target
        remaining = summary.remaining
        summary_sheet.append(
            [
                day.isoformat(),
                summary.total.calories,
                summary.total.protein,
                summary.total.fat,
                summary.total.carbs,
                target.calories if target else None,
                target.protein if target else None,
                target.fat if target else None,
                target.carbs if target else None,
                remaining.calories if remaining else None,
                remaining.protein if remaining else None,
                remaining.fat if remaining else None,
                remaining.carbs if remaining else None,
            ]
        )

    settings_sheet = workbook.create_sheet("settings")
    settings_sheet.append(["field", "value"])
    if settings is not None:
        settings_sheet.append(["height_cm", settings.height_cm])
        settings_sheet.append(["weight_kg", settings.weight_kg])
        settings_sheet.append(["target_calories", settings.target_calories])
        settings_sheet.append(["target_protein", settings.target_protein])
        settings_sheet.append(["target_fat", settings.target_fat])
        settings_sheet.append(["target_carbs", settings.target_carbs])
        settings_sheet.append(["updated_at", settings.updated_at.isoformat(timespec="seconds")])

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 32)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_filename(start_date: date | None, end_date: date) -> str:
    if start_date is None:
        return f"kbju_all_until_{end_date.isoformat()}.xlsx"
    if start_date == end_date:
        return f"kbju_{end_date.isoformat()}.xlsx"
    return f"kbju_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"


def export_caption(entries_count: int) -> str:
    return f"📤 Выгрузка готова. Записей: {format_number(entries_count)}"
